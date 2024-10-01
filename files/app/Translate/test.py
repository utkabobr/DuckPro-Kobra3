import openpyxl
import os
import xml.etree.ElementTree as ET

# 当前文件地址
base_dir = os.path.dirname(os.path.abspath(__file__))  # 'G:\\site\\python\\learn\\base'
book_name_xlsx = os.path.join(base_dir, 'K3系列直驱屏UI文案.xlsx')  # 'G:\\site\\python\\learn\\base\\xlsx格式测试工作簿.xlsx'

sheet_name_xlsx = '工作表1'

langTolang = {
    "英文文案": "LanguageEnglish", 
    "德语文案": "LanguageGerman", 
    "法语文案": "LanguageFrench", 
    "意大利文案": "LanguageItaly", 
    "西班牙语": "LanguageSpain", 
    "俄语": "LanguageRussian", 
    "日语文案": "LanguageJapanese"
}

Language = [
    "简体中文",
    "English",
    "Deutsch",
    "Français",
    "Italiano",
    "Español",
    "Русский",
    "日本語"
]


def write_xml_message(root, source_text, translation_text):
    source_text = source_text.strip()
    translation_text = translation_text.strip()
    message = ET.SubElement(root, "message")
    source = ET.SubElement(message, "source")
    source.text = source_text
    translation = ET.SubElement(message, "translation")
    translation.text = translation_text
    return message

def read_excel_xlsx_1(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    langs = []
    lang_info = []
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=6):
        for cell in row:
            if cell.value:
                langs.append(langTolang[cell.value])
    print(langTolang)
    row_index = -1    
    for row in sheet.iter_rows(min_row=3, min_col=6):
        row_index = row_index + 1
        lang_info_row = {}
        cell_index = -1
        for cell in row:
            if cell.value: 
                cell_index =cell_index + 1
                lang_info_row[langs[cell_index]] = cell.value.split("\n")
        lang_info.append(lang_info_row)

    xml_info = {}
    for lang in langs:
        root = ET.Element("TS")
        root.set("version", "2.1")
        root.set("language", "zh_CN")
        context = ET.SubElement(root, "context")
        name = ET.SubElement(context, "name")
        name.text = "MainWindow"
        message = write_xml_message(context, "MainWindow", "")
        location = ET.SubElement(message, "location")
        location.set("filename", "../mainwindow.ui")
        location.set("line", "14")
        xml_info[lang] = root
    
    for lang_info_row in lang_info:
        for lang, infos in lang_info_row.items():
            if lang != 'LanguageEnglish':
                en_infos = lang_info_row['LanguageEnglish']
                for i in range(len(infos)):
                    if i < len(en_infos):
                        en_infos[i] = en_infos[i].strip()
                        infos[i] = infos[i].strip()
                        if en_infos[i] != "" and not en_infos[i].startswith("CODE:") and Language.count(en_infos[i]) == 0 and infos[i] != "":
                            write_xml_message(xml_info[lang].find("context"), en_infos[i], infos[i])
                # 创建ElementTree对象
                tree = ET.ElementTree(xml_info[lang])
                ET.indent(tree)
                tree.write(lang + ".ts", encoding="utf-8", method='xml', xml_declaration=True)

read_excel_xlsx_1(book_name_xlsx, sheet_name_xlsx)
